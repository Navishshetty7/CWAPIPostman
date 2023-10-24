import requests
import openpyxl

url_template = "https://prod-v2-api.curbwaste.com/api/transactions/sync/{}"

workbook = openpyxl.Workbook()
sheet = workbook.active
sheet["A1"] = "Value"
sheet["B1"] = "qbid"
url_last_workbook = openpyxl.load_workbook("url_last.xlsx")
url_last_sheet = url_last_workbook.active

for row, row_data in enumerate(url_last_sheet.iter_rows(values_only=True), start=2):
    value = row_data[0]
    print(value)
    url = url_template.format(value)
    response = requests.post(url)
    data = response.json()
    print(data)
    qbId = data.get('data', {}).get('qbId')
    print(qbId)
    sheet[f"A{row}"] = value
    sheet[f"B{row}"] = qbId

workbook.save("output.xlsx")
workbook.close()
url_last_workbook.close()
